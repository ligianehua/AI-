"""线索业务逻辑：CRUD / 查重 / Excel 导入 / 转化 / 分配。RBAC 在本层强制。"""

import io
import re
import uuid
from datetime import UTC, datetime
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import Select, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ConflictError, DomainError, PermissionDeniedError
from app.models.account import Account
from app.models.activity import Activity
from app.models.contact import Contact
from app.models.enums import (
    ActivityRelatedType,
    LeadSource,
    LeadStatus,
    OpportunityStage,
    Role,
)
from app.models.lead import Lead
from app.models.opportunity import Opportunity
from app.models.user import User
from app.schemas.lead import (
    AssignResult,
    ConvertRequest,
    ConvertResult,
    DuplicateWarning,
    ImportRowError,
    LeadCreate,
    LeadImportReport,
    LeadUpdate,
)
from app.services.base import BaseService
from app.tasks import dispatcher

PHONE_RE = re.compile(r"^1\d{10}$")

SOURCE_ALIASES: dict[str, LeadSource] = {
    "官网": LeadSource.WEBSITE,
    "官网咨询": LeadSource.WEBSITE,
    "展会": LeadSource.EXHIBITION,
    "转介绍": LeadSource.REFERRAL,
    "广告": LeadSource.ADS,
    "陌拜": LeadSource.COLD_CALL,
    "其他": LeadSource.OTHER,
    **{s.value: s for s in LeadSource},
}

IMPORT_COLUMNS = ["客户公司*", "联系人", "手机号", "微信", "来源*", "行业", "需求描述"]


class LeadService(BaseService[Lead]):
    model = Lead
    sortable_fields = frozenset({"created_at", "score", "status", "account_name"})


lead_service = LeadService()


async def find_duplicates(
    session: AsyncSession,
    contact_phone: str | None,
    account_name: str | None,
    exclude_id: uuid.UUID | None = None,
) -> list[DuplicateWarning]:
    """撞单检测：全量查（跨可见域是业务要求），只返回最小信息，不硬拦截。"""
    conditions = []
    if contact_phone:
        conditions.append(Lead.contact_phone == contact_phone)
    if account_name:
        conditions.append(Lead.account_name == account_name)
    if not conditions:
        return []
    stmt = (
        select(Lead, User.name)
        .join(User, User.id == Lead.owner_id)
        .where(
            Lead.deleted_at.is_(None),
            Lead.status != LeadStatus.INVALID,
            conditions[0] if len(conditions) == 1 else (conditions[0] | conditions[1]),
        )
        .limit(5)
    )
    if exclude_id:
        stmt = stmt.where(Lead.id != exclude_id)
    rows = (await session.execute(stmt)).all()
    return [
        DuplicateWarning(
            lead_id=lead.id,
            account_name=lead.account_name,
            owner_name=owner_name,
            matched_field=(
                "contact_phone"
                if contact_phone and lead.contact_phone == contact_phone
                else "account_name"
            ),
        )
        for lead, owner_name in rows
    ]


async def create_lead(
    session: AsyncSession, actor: User, payload: LeadCreate
) -> tuple[Lead, list[DuplicateWarning]]:
    lead = Lead(**payload.model_dump(), owner_id=actor.id, status=LeadStatus.NEW)
    session.add(lead)
    await session.commit()
    await session.refresh(lead)
    warnings = await find_duplicates(
        session, lead.contact_phone, lead.account_name, exclude_id=lead.id
    )
    await dispatcher.enqueue("score_lead_task", str(lead.id))
    return lead, warnings


def _list_query(
    actor: User,
    service: LeadService,
    status: LeadStatus | None,
    score_gte: int | None,
    owner_id: uuid.UUID | None,
) -> Select[tuple[Lead]]:
    stmt = service.base_query(actor)
    if status is not None:
        stmt = stmt.where(Lead.status == status)
    if score_gte is not None:
        stmt = stmt.where(Lead.score >= score_gte)
    if owner_id is not None:
        stmt = stmt.where(Lead.owner_id == owner_id)
    return stmt


async def list_leads(
    session: AsyncSession,
    actor: User,
    *,
    status: LeadStatus | None = None,
    score_gte: int | None = None,
    owner_id: uuid.UUID | None = None,
    page: int = 1,
    page_size: int = 20,
    sort: str | None = "-score",
) -> tuple[list[tuple[Lead, str]], int]:
    """返回 [(lead, owner_name)], total。"""
    stmt = _list_query(actor, lead_service, status, score_gte, owner_id)
    count_stmt = select(func.count()).select_from(stmt.order_by(None).subquery())
    total = int(await session.scalar(count_stmt) or 0)

    sorted_stmt = lead_service.apply_sort(stmt, sort)
    if sort in ("-score", "score"):  # 未评分的排最后
        sorted_stmt = stmt.order_by(
            Lead.score.desc().nulls_last() if sort == "-score" else Lead.score.asc().nulls_last()
        )
    rows = (
        await session.execute(
            sorted_stmt.add_columns(User.name)
            .join(User, User.id == Lead.owner_id)
            .limit(page_size)
            .offset((page - 1) * page_size)
        )
    ).all()
    return [(lead, owner_name) for lead, owner_name in rows], total


async def get_lead_detail(
    session: AsyncSession, actor: User, lead_id: uuid.UUID
) -> tuple[Lead, str, list[Activity]]:
    lead = await lead_service.get(session, actor, lead_id)
    owner_name = await session.scalar(select(User.name).where(User.id == lead.owner_id))
    activities = list(
        await session.scalars(
            select(Activity)
            .where(
                Activity.related_type == ActivityRelatedType.LEAD,
                Activity.related_id == lead.id,
                Activity.deleted_at.is_(None),
            )
            .order_by(Activity.created_at.desc())
        )
    )
    return lead, owner_name or "", activities


async def update_lead(
    session: AsyncSession, actor: User, lead_id: uuid.UUID, payload: LeadUpdate
) -> Lead:
    lead = await lead_service.get(session, actor, lead_id)
    data = payload.model_dump(exclude_unset=True)
    if data.get("status") == LeadStatus.CONVERTED:
        raise DomainError("转化状态只能通过转化操作设置")
    if lead.status == LeadStatus.CONVERTED and "status" in data:
        raise DomainError("已转化线索不能改状态")
    for key, value in data.items():
        setattr(lead, key, value)
    await session.commit()
    await session.refresh(lead)
    return lead


async def trigger_rescore(session: AsyncSession, actor: User, lead_id: uuid.UUID) -> None:
    lead = await lead_service.get(session, actor, lead_id)
    await dispatcher.enqueue("score_lead_task", str(lead.id))


async def convert_lead(
    session: AsyncSession, actor: User, lead_id: uuid.UUID, payload: ConvertRequest
) -> ConvertResult:
    """转化：一个事务里创建 account + contact + opportunity 并回链线索，失败整体回滚。"""
    lead = await lead_service.get(session, actor, lead_id)
    if lead.status == LeadStatus.CONVERTED:
        raise ConflictError("该线索已转化")
    if lead.status == LeadStatus.INVALID:
        raise DomainError("无效线索不能转化")

    account = Account(
        name=payload.account_name or lead.account_name,
        industry=lead.industry,
        owner_id=lead.owner_id,
    )
    session.add(account)
    await session.flush()

    contact: Contact | None = None
    if lead.contact_name or lead.contact_phone:
        contact = Contact(
            account_id=account.id,
            name=lead.contact_name or "未知联系人",
            phone=lead.contact_phone,
            wechat=lead.contact_wechat,
        )
        session.add(contact)
        await session.flush()

    opportunity = Opportunity(
        account_id=account.id,
        name=payload.opportunity_name or f"{account.name}-初始商机",
        amount=payload.amount if payload.amount is not None else 0,
        stage=OpportunityStage.INITIAL,
        owner_id=lead.owner_id,
        stage_history=[
            {
                "stage": OpportunityStage.INITIAL.value,
                "entered_at": datetime.now(UTC).isoformat(),
                "by": actor.name,
            }
        ],
    )
    session.add(opportunity)
    await session.flush()

    lead.status = LeadStatus.CONVERTED
    lead.converted_account_id = account.id
    lead.converted_opportunity_id = opportunity.id
    await session.commit()

    return ConvertResult(
        account_id=account.id,
        contact_id=contact.id if contact else None,
        opportunity_id=opportunity.id,
    )


async def assign_leads(
    session: AsyncSession, actor: User, lead_ids: list[uuid.UUID], owner_id: uuid.UUID
) -> AssignResult:
    """批量分配（manager 限本团队目标；admin 不限；sales 禁止）。"""
    if actor.role == Role.SALES:
        raise PermissionDeniedError("销售不能分配线索")
    target = await session.scalar(
        select(User).where(User.id == owner_id, User.deleted_at.is_(None), User.is_active)
    )
    if target is None:
        raise DomainError("目标负责人不存在或已停用")
    if actor.role == Role.MANAGER and target.team_id != actor.team_id:
        raise DomainError("只能分配给本团队成员")

    leads = list(await session.scalars(lead_service.base_query(actor).where(Lead.id.in_(lead_ids))))
    if len(leads) != len(set(lead_ids)):
        raise DomainError("部分线索不存在或无权操作")
    for lead in leads:
        lead.owner_id = owner_id
    await session.commit()
    return AssignResult(assigned=len(leads))


# ---------- Excel 导入 ----------


def build_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "线索导入"
    ws.append(IMPORT_COLUMNS)
    ws.append(
        [
            "杭州示例科技",
            "王先生",
            "13800138000",
            "wx_demo",
            "转介绍",
            "制造业",
            "需要一套 CRM，预算 20 万",
        ]
    )
    for idx in range(1, len(IMPORT_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_row(values: tuple[Any, ...]) -> LeadCreate:
    def text(idx: int) -> str | None:
        if idx >= len(values) or values[idx] is None:
            return None
        s = str(values[idx]).strip()
        return s or None

    account_name = text(0)
    if not account_name:
        raise ValueError("客户公司不能为空")
    source_raw = text(4)
    if not source_raw:
        raise ValueError("来源不能为空")
    source = SOURCE_ALIASES.get(source_raw)
    if source is None:
        raise ValueError(f"来源不合法：{source_raw}（可选：官网/展会/转介绍/广告/陌拜/其他）")
    phone = text(2)
    if phone and not PHONE_RE.match(phone):
        raise ValueError(f"手机号格式错误：{phone}")
    return LeadCreate(
        source=source,
        account_name=account_name,
        contact_name=text(1),
        contact_phone=phone,
        contact_wechat=text(3),
        industry=text(5),
        requirement_desc=text(6),
    )


async def import_leads(session: AsyncSession, actor: User, file: BinaryIO) -> LeadImportReport:
    """Excel 导入：逐行校验、一次落库、行号级错误报告、撞单提示（不拦截）。"""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        raise DomainError("无法读取 Excel 文件，请使用模板（.xlsx）") from exc
    ws = wb.active
    if ws is None:
        raise DomainError("Excel 文件没有工作表")

    rows = ws.iter_rows(min_row=2, values_only=True)
    parsed: list[tuple[int, LeadCreate]] = []
    errors: list[ImportRowError] = []
    total = 0
    for row_no, values in enumerate(rows, start=2):
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        total += 1
        try:
            parsed.append((row_no, _parse_row(values)))
        except Exception as exc:  # ValueError / pydantic ValidationError 都归错误行
            errors.append(ImportRowError(row=row_no, reason=str(exc)))
    wb.close()
    if total > 2000:
        raise DomainError("单次导入最多 2000 行")

    # 撞单提示：与库内已有线索按手机号/公司名比对（一次性查询）
    phones = {p.contact_phone for _, p in parsed if p.contact_phone}
    names = {p.account_name for _, p in parsed}
    existing_phones: set[str] = set()
    existing_names: set[str] = set()
    if phones:
        existing_phones = {
            p
            for p in await session.scalars(
                select(Lead.contact_phone).where(
                    Lead.contact_phone.in_(phones), Lead.deleted_at.is_(None)
                )
            )
            if p
        }
    if names:
        existing_names = set(
            await session.scalars(
                select(Lead.account_name).where(
                    Lead.account_name.in_(names), Lead.deleted_at.is_(None)
                )
            )
        )
    duplicate_warnings = [
        ImportRowError(
            row=row_no,
            reason=(
                f"疑似撞单：手机号 {p.contact_phone} 已存在"
                if p.contact_phone in existing_phones
                else f"疑似撞单：公司「{p.account_name}」已存在"
            ),
        )
        for row_no, p in parsed
        if p.contact_phone in existing_phones or p.account_name in existing_names
    ]

    leads = [Lead(**p.model_dump(), owner_id=actor.id, status=LeadStatus.NEW) for _, p in parsed]
    session.add_all(leads)
    await session.commit()
    for lead in leads:
        await dispatcher.enqueue("score_lead_task", str(lead.id))

    return LeadImportReport(
        total_rows=total,
        imported=len(leads),
        failed=len(errors),
        errors=errors,
        duplicate_warnings=duplicate_warnings,
    )
