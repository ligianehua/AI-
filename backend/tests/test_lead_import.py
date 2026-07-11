"""Excel 导入测试：模板、错误行报告、撞单提示、500 行性能。"""

import io
import time
from collections.abc import Awaitable, Callable
from typing import Any

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.lead import Lead
from app.services.lead_service import IMPORT_COLUMNS
from app.tasks import dispatcher
from tests.conftest import RoleUsers

LoginFn = Callable[[str], Awaitable[dict[str, str]]]
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(autouse=True)
def enqueued(monkeypatch: pytest.MonkeyPatch) -> list[tuple[str, tuple[Any, ...]]]:
    calls: list[tuple[str, tuple[Any, ...]]] = []

    async def fake_enqueue(task_name: str, *args: Any) -> None:
        calls.append((task_name, args))

    monkeypatch.setattr(dispatcher, "enqueue", fake_enqueue)
    return calls


def _xlsx(rows: list[list[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(IMPORT_COLUMNS)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _upload(client: AsyncClient, headers: dict[str, str], content: bytes) -> Any:
    return await client.post(
        "/api/v1/leads/import",
        files={"file": ("leads.xlsx", content, XLSX_MIME)},
        headers=headers,
    )


async def test_download_template(client: AsyncClient, roles: RoleUsers, login: LoginFn) -> None:
    headers = await login("sales_a@test.cn")
    resp = await client.get("/api/v1/leads/import-template", headers=headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(XLSX_MIME)
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws is not None
    assert [c.value for c in ws[1]] == IMPORT_COLUMNS


async def test_import_with_error_rows(
    client: AsyncClient,
    session: AsyncSession,
    roles: RoleUsers,
    login: LoginFn,
    enqueued: list[tuple[str, tuple[Any, ...]]],
) -> None:
    content = _xlsx(
        [
            ["北京优选科技", "赵总", "13700000001", "wx_a", "展会", "互联网", "要 CRM，预算 10 万"],
            [None, "无公司", "13700000002", None, "展会", None, None],  # 行3：缺公司
            ["广州测试公司", None, "123", None, "官网", None, None],  # 行4：手机号错
            ["深圳测试公司", None, None, None, "不存在的来源", None, None],  # 行5：来源错
            ["成都正常公司", "钱总", None, None, "转介绍", "教育", None],  # 行6：正常
        ]
    )
    headers = await login("sales_a@test.cn")
    resp = await _upload(client, headers, content)
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["total_rows"] == 5
    assert report["imported"] == 2
    assert report["failed"] == 3
    assert {e["row"] for e in report["errors"]} == {3, 4, 5}
    reasons = {e["row"]: e["reason"] for e in report["errors"]}
    assert "公司" in reasons[3]
    assert "手机号" in reasons[4]
    assert "来源" in reasons[5]

    count = await session.scalar(
        select(func.count()).select_from(Lead).where(Lead.owner_id == roles.sales_a.id)
    )
    assert count == 2
    assert len(enqueued) == 2  # 每条导入的线索都触发评分


async def test_import_duplicate_warning(
    client: AsyncClient, session: AsyncSession, roles: RoleUsers, login: LoginFn
) -> None:
    session.add(
        Lead(
            source="website",
            account_name="已有公司",
            contact_phone="13700009999",
            status="new",
            owner_id=roles.sales_b.id,
        )
    )
    await session.commit()

    content = _xlsx([["已有公司", None, "13700009999", None, "官网", None, None]])
    resp = await _upload(client, await login("sales_a@test.cn"), content)
    report = resp.json()
    assert report["imported"] == 1  # 提示不拦截
    assert len(report["duplicate_warnings"]) == 1
    assert "撞单" in report["duplicate_warnings"][0]["reason"]


async def test_import_500_rows_under_10s(
    client: AsyncClient, session: AsyncSession, roles: RoleUsers, login: LoginFn
) -> None:
    rows = [
        [
            f"批量公司{i}",
            f"联系人{i}",
            f"137{i:08d}",
            None,
            "官网",
            "制造业",
            "需要销售管理工具，尽快对接",
        ]
        for i in range(500)
    ]
    content = _xlsx(rows)
    headers = await login("sales_a@test.cn")
    started = time.perf_counter()
    resp = await _upload(client, headers, content)
    elapsed = time.perf_counter() - started
    assert resp.status_code == 200
    assert resp.json()["imported"] == 500
    assert elapsed < 10, f"导入耗时 {elapsed:.2f}s 超过 10s"


async def test_import_garbage_file_rejected(
    client: AsyncClient, roles: RoleUsers, login: LoginFn
) -> None:
    resp = await _upload(client, await login("sales_a@test.cn"), b"this is not xlsx")
    assert resp.status_code == 400
    assert "Excel" in resp.json()["message"]
